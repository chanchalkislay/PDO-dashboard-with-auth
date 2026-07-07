"""Tab 14 — Finder & Reports.

Three independent sections driven by the sidebar's product / CY year / period:

  ① RO / TA Finder        — ranked query builder with independent scope filters
  ② Performance YoY       — multi-level bulk xlsx (DO / RSA / District / COM / Highway)
  ③ TA Profile Bulk       — multiple TA profile sheets in one xlsx

Market-share denominator is user-selectable (Industry or PSU) in each section.
All scope filters are independent across sections and do not affect the sidebar.
"""
from __future__ import annotations
import io
import streamlit as st
from components.downloads import df_download
import pandas as pd

from core import (
    share_frame, totals_row, ta_volume_grid,
    OMC_ORDER, PSU, PVT, COM_LABELS,
    indian, pct, pp, klpm,
)

# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────
_METRICS = [
    "Volume CY (KL)",
    "Volume LY (KL)",
    "Volume change (KL)",
    "Volume growth %",
    "Market share CY %",
    "Market share LY %",
    "Share change (pp)",
    "Notional gain / loss (KL)",
    "KLPM",
]
_METRIC_COL = {
    "Volume CY (KL)":            "vol_cy",
    "Volume LY (KL)":            "vol_ly",
    "Volume change (KL)":        "vol_chg",
    "Volume growth %":           "vol_gr",
    "Market share CY %":         "shr_cy",
    "Market share LY %":         "shr_ly",
    "Share change (pp)":         "shr_pp",
    "Notional gain / loss (KL)": "notional",
    "KLPM":                      "klpm_v",
}


# ─────────────────────────────────────────────────────────────────────────────
# Shared helpers
# ─────────────────────────────────────────────────────────────────────────────
def _apply_scope(df, f_rsa, f_dist, f_com, f_hwytype, f_hwyno):
    """Apply independent scope filters to any frame that has the standard columns."""
    m = pd.Series(True, index=df.index)
    if f_rsa:     m &= df["rsa_code"].isin(f_rsa)
    if f_dist:    m &= df["district"].isin(f_dist)
    if f_com:     m &= df["com"].isin(f_com)
    if f_hwytype: m &= df["hwy_type"].isin(f_hwytype)
    if f_hwyno:   m &= df["highway_no"].isin(f_hwyno)
    return df[m]


def _scope_widgets(prefix, ctx):
    """Render 5 independent scope filter multiselects; return filter tuple."""
    rsa_labels = ctx["rsa_labels"]
    c1, c2, c3, c4, c5 = st.columns(5)
    f_rsa_lbl = c1.multiselect("RSA",          list(rsa_labels), key=f"{prefix}_rsa")
    f_dist    = c2.multiselect("District",      ctx["districts"], key=f"{prefix}_dist")
    f_com     = c3.multiselect("COM",           ctx["coms"],
                                format_func=lambda c: COM_LABELS[c],
                                key=f"{prefix}_com")
    f_hwytype = c4.multiselect("Highway type",
                                ["NH", "SH", "Non-Highway"], key=f"{prefix}_hwytype")
    f_hwyno   = c5.multiselect("Highway no.",  ctx["hwy_nums"], key=f"{prefix}_hwyno")
    return (
        [rsa_labels[l] for l in f_rsa_lbl],
        f_dist, f_com, f_hwytype, f_hwyno,
    )


# ─────────────────────────────────────────────────────────────────────────────
# ① RO / TA Finder
# ─────────────────────────────────────────────────────────────────────────────
@st.fragment
def _render_finder(ctx):
    monthly    = ctx["monthly"]
    ro_master  = ctx["ro_master"]
    cy         = ctx["cy"]
    ly         = ctx["ly"]
    product    = ctx["product"]
    months     = ctx["months"]
    period_lbl = ctx["period_lbl"]
    n_months   = max(len(months), 1)
    TA_NAME    = ctx["TA_NAME"]

    # ── Query builder controls ────────────────────────────────────────────────
    st.markdown("##### Query builder")
    qc1, qc2, qc3, qc4 = st.columns([2, 3, 2, 3])
    entity     = qc1.radio("Entity",    ["Retail Outlet (RO)", "Trading Area (TA)"],
                           key="fn_entity", horizontal=True)
    sel_metric = qc2.selectbox("Rank by", _METRICS, key="fn_metric")
    direction  = qc3.radio("Direction", ["Highest first", "Lowest first"],
                           key="fn_dir", horizontal=True)
    sel_n      = qc4.select_slider("Show top N",
                                   [3, 5, 10, 20, 50, "All"],
                                   value=10, key="fn_n")

    sc1, sc2 = st.columns([3, 2])
    sel_omcs = sc1.multiselect("OMC(s)", OMC_ORDER, default=["IOCL"], key="fn_omc")
    if not sel_omcs:
        sel_omcs = ["IOCL"]
    denom    = sc2.radio("Share denominator", ["Industry", "PSU"],
                         horizontal=True, key="fn_denom",
                         help="Denominator for RO/TA market share within the Trading Area.")
    univ     = OMC_ORDER if denom == "Industry" else PSU

    st.caption("Scope filters — independent of sidebar (empty = whole DO)")
    f_rsa, f_dist, f_com, f_hwytype, f_hwyno = _scope_widgets("fn", ctx)
    sort_col  = _METRIC_COL[sel_metric]
    ascending = (direction == "Lowest first")

    # Period-filtered full monthly data (NOT scope-filtered — so TA totals are complete)
    cy_all = monthly[
        (monthly["fy_code"] == cy) &
        (monthly["product"] == product) &
        (monthly["month_index"].isin(months))
    ]
    ly_all = (monthly[
        (monthly["fy_code"] == ly) &
        (monthly["product"] == product) &
        (monthly["month_index"].isin(months))
    ] if ly else monthly.iloc[0:0])

    # ── RO mode ───────────────────────────────────────────────────────────────
    if entity == "Retail Outlet (RO)":
        finder_ro = _apply_scope(
            ro_master[ro_master["omc"].isin(sel_omcs)],
            f_rsa, f_dist, f_com, f_hwytype, f_hwyno)
        scope_saps = set(finder_ro["sap_code"].astype(str))
        if not scope_saps:
            st.info("No ROs match the selected scope filters.")
            return

        # Per-RO volumes
        cy_omc = (cy_all[cy_all["omc"].isin(sel_omcs)]
                  .groupby("sap_code")["volume_kl"].sum())
        ly_omc = (ly_all[ly_all["omc"].isin(sel_omcs)]
                  .groupby("sap_code")["volume_kl"].sum()
                  .reindex(cy_omc.index, fill_value=0.0))

        # TA totals for market share (use full monthly — not scoped — for correct denominator)
        sap_ta    = ro_master.set_index("sap_code")["ta_code"].to_dict()
        cy_ta_tot = (cy_all[cy_all["omc"].isin(univ)]
                     .groupby("ta_code")["volume_kl"].sum())
        ly_ta_tot = (ly_all[ly_all["omc"].isin(univ)]
                     .groupby("ta_code")["volume_kl"].sum())

        ro_df = pd.DataFrame({"vol_cy": cy_omc,
                               "vol_ly": ly_omc}).fillna(0.0)
        ro_df["vol_chg"] = ro_df["vol_cy"] - ro_df["vol_ly"]
        ro_df["vol_gr"]  = (ro_df["vol_chg"] / ro_df["vol_ly"] * 100
                            ).where(ro_df["vol_ly"] > 0)
        ro_df["ta_code"] = ro_df.index.map(sap_ta)
        ta_cy_s          = ro_df["ta_code"].map(cy_ta_tot)
        ta_ly_s          = ro_df["ta_code"].map(ly_ta_tot)
        ro_df["shr_cy"]  = (ro_df["vol_cy"] / ta_cy_s * 100).where(ta_cy_s > 0)
        ro_df["shr_ly"]  = (ro_df["vol_ly"] / ta_ly_s * 100).where(ta_ly_s > 0)
        ro_df["shr_pp"]  = (ro_df["shr_cy"] - ro_df["shr_ly"]
                            ).where(ro_df["shr_cy"].notna() & ro_df["shr_ly"].notna())
        ro_df["notional"]= (ro_df["shr_pp"] / 100 * ta_cy_s
                            ).where(ro_df["shr_pp"].notna())
        ro_df["klpm_v"]  = ro_df["vol_cy"] / n_months

        # Apply scope, sort, top-N
        ro_df = ro_df[ro_df.index.isin(scope_saps)]
        ro_df = ro_df.sort_values(sort_col, ascending=ascending, na_position="last")
        if sel_n != "All":
            ro_df = ro_df.head(int(sel_n))
        if ro_df.empty:
            st.info("No data for the selected filters and period.")
            return

        # Attach display metadata
        meta = (ro_master[ro_master["sap_code"].isin(ro_df.index)]
                [["sap_code","ro_name","omc","rsa_name","district",
                  "com","ta_code","highway_no"]]
                .drop_duplicates("sap_code").set_index("sap_code"))
        ro_df = ro_df.join(meta, how="left", rsuffix="_m")
        if "ta_code_m" in ro_df.columns:
            ro_df["ta_code"] = ro_df["ta_code_m"].fillna(ro_df["ta_code"])
        ro_df["ta_name"] = ro_df["ta_code"].map(TA_NAME).fillna("")

        n_shown = len(ro_df)
        st.caption(
            f"**{direction[:7]} {n_shown} RO(s)** by **{sel_metric}** · "
            f"OMC: {', '.join(sel_omcs)} · {product} · {period_lbl} · "
            f"{cy} vs {ly or '—'} · share within TA ({denom} denominator)")

        disp = pd.DataFrame({
            "Rank":     range(1, n_shown + 1),
            "SAP Code": ro_df.index.tolist(),
            "RO Name":  ro_df["ro_name"].fillna("—").tolist(),
            "OMC":      ro_df["omc"].fillna("—").tolist(),
            "RSA":      ro_df["rsa_name"].fillna("—").tolist(),
            "District": ro_df["district"].fillna("—").tolist(),
            "COM":      ro_df["com"].fillna("—").tolist(),
            "TA":       ro_df["ta_name"].tolist(),
            "Highway":  ro_df.get("highway_no", pd.Series(dtype=str)).fillna("").tolist()
                        if "highway_no" in ro_df.columns else [""] * n_shown,
            "Vol CY (KL)":              [round(v, 1) for v in ro_df["vol_cy"]],
            "Vol LY (KL)":              [round(v, 1) for v in ro_df["vol_ly"]],
            "Change (KL)":              [round(v, 1) for v in ro_df["vol_chg"]],
            "Growth %":                 ro_df["vol_gr"].apply(
                lambda x: f"{x:.2f}%" if pd.notna(x) else "—").tolist(),
            f"TA Shr CY% ({denom})":    ro_df["shr_cy"].apply(
                lambda x: f"{x:.2f}%" if pd.notna(x) else "—").tolist(),
            f"TA Shr LY% ({denom})":    ro_df["shr_ly"].apply(
                lambda x: f"{x:.2f}%" if pd.notna(x) else "—").tolist(),
            "Share pp":                 ro_df["shr_pp"].apply(
                lambda x: f"{x:+.2f}" if pd.notna(x) else "—").tolist(),
            "Notional (KL)":            ro_df["notional"].apply(
                lambda x: round(x, 1) if pd.notna(x) else "—").tolist(),
            "KLPM":                     [round(v, 2) for v in ro_df["klpm_v"]],
        })
        st.dataframe(disp, hide_index=True, use_container_width=True)
        df_download(disp, "t14_1")

        buf = io.BytesIO()
        disp.to_excel(buf, index=False, engine="openpyxl", sheet_name="RO_Finder")
        fname = (f"RO_Finder_{'+'.join(sel_omcs)}_{sel_metric.split()[0]}_"
                 f"{product}_{cy}_{period_lbl}.xlsx"
                 ).replace(" ", "_").replace("/", "").replace("(", "").replace(")", "")
        st.download_button("⬇ Download results as Excel (.xlsx)",
                           data=buf.getvalue(), file_name=fname,
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="fn_ro_dl")

    # ── TA mode ───────────────────────────────────────────────────────────────
    else:
        finder_m = _apply_scope(monthly, f_rsa, f_dist, f_com, f_hwytype, f_hwyno)
        cy_s = finder_m[(finder_m["fy_code"] == cy) &
                        (finder_m["product"] == product) &
                        (finder_m["month_index"].isin(months))]
        ly_s = (finder_m[(finder_m["fy_code"] == ly) &
                         (finder_m["product"] == product) &
                         (finder_m["month_index"].isin(months))]
                if ly else finder_m.iloc[0:0])

        cy_omc_ta  = (cy_s[cy_s["omc"].isin(sel_omcs)]
                      .groupby("ta_code")["volume_kl"].sum())
        ly_omc_ta  = (ly_s[ly_s["omc"].isin(sel_omcs)]
                      .groupby("ta_code")["volume_kl"].sum())
        cy_univ_ta = (cy_s[cy_s["omc"].isin(univ)]
                      .groupby("ta_code")["volume_kl"].sum())
        ly_univ_ta = (ly_s[ly_s["omc"].isin(univ)]
                      .groupby("ta_code")["volume_kl"].sum())

        ta_df = pd.DataFrame({"vol_cy": cy_omc_ta}).fillna(0.0)
        ta_df["vol_ly"]  = ly_omc_ta.reindex(ta_df.index, fill_value=0.0)
        ta_df["vol_chg"] = ta_df["vol_cy"] - ta_df["vol_ly"]
        ta_df["vol_gr"]  = (ta_df["vol_chg"] / ta_df["vol_ly"] * 100
                            ).where(ta_df["vol_ly"] > 0)
        ta_tot_cy        = cy_univ_ta.reindex(ta_df.index, fill_value=0.0)
        ta_tot_ly        = ly_univ_ta.reindex(ta_df.index, fill_value=0.0)
        ta_df["shr_cy"]  = (ta_df["vol_cy"] / ta_tot_cy * 100).where(ta_tot_cy > 0)
        ta_df["shr_ly"]  = (ta_df["vol_ly"] / ta_tot_ly * 100).where(ta_tot_ly > 0)
        ta_df["shr_pp"]  = (ta_df["shr_cy"] - ta_df["shr_ly"]
                            ).where(ta_df["shr_cy"].notna() & ta_df["shr_ly"].notna())
        ta_df["notional"]= (ta_df["shr_pp"] / 100 * ta_tot_cy
                            ).where(ta_df["shr_pp"].notna())
        ta_ro_cnt        = (_apply_scope(ro_master[ro_master["omc"].isin(sel_omcs)],
                                         f_rsa, f_dist, f_com, f_hwytype, f_hwyno)
                            .groupby("ta_code")["sap_code"].nunique()
                            .reindex(ta_df.index, fill_value=1))
        ta_df["klpm_v"]  = (ta_df["vol_cy"] / n_months / ta_ro_cnt
                            ).where(ta_ro_cnt > 0)
        ta_df["ro_cnt"]  = ta_ro_cnt

        ta_df = ta_df.sort_values(sort_col, ascending=ascending, na_position="last")
        if sel_n != "All":
            ta_df = ta_df.head(int(sel_n))
        if ta_df.empty:
            st.info("No TAs match the selected scope filters.")
            return

        ta_df["ta_name"] = ta_df.index.map(TA_NAME).fillna("")
        ta_meta = (ro_master.drop_duplicates("ta_code")
                   .set_index("ta_code")[["rsa_name", "district"]])
        ta_df = ta_df.join(ta_meta, how="left")

        n_shown = len(ta_df)
        st.caption(
            f"**{direction[:7]} {n_shown} TA(s)** by **{sel_metric}** · "
            f"OMC: {', '.join(sel_omcs)} · {product} · {period_lbl} · "
            f"share denominator: {denom}")

        disp = pd.DataFrame({
            "Rank":     range(1, n_shown + 1),
            "TA Code":  ta_df.index.tolist(),
            "TA Name":  ta_df["ta_name"].tolist(),
            "RSA":      ta_df["rsa_name"].fillna("—").tolist(),
            "District": ta_df["district"].fillna("—").tolist(),
            "RO Count": ta_df["ro_cnt"].tolist(),
            "Vol CY (KL)":           [round(v, 1) for v in ta_df["vol_cy"]],
            "Vol LY (KL)":           [round(v, 1) for v in ta_df["vol_ly"]],
            "Change (KL)":           [round(v, 1) for v in ta_df["vol_chg"]],
            "Growth %":              ta_df["vol_gr"].apply(
                lambda x: f"{x:.2f}%" if pd.notna(x) else "—").tolist(),
            f"Shr CY% ({denom})":    ta_df["shr_cy"].apply(
                lambda x: f"{x:.2f}%" if pd.notna(x) else "—").tolist(),
            f"Shr LY% ({denom})":    ta_df["shr_ly"].apply(
                lambda x: f"{x:.2f}%" if pd.notna(x) else "—").tolist(),
            "Share pp":              ta_df["shr_pp"].apply(
                lambda x: f"{x:+.2f}" if pd.notna(x) else "—").tolist(),
            "Notional (KL)":         ta_df["notional"].apply(
                lambda x: round(x, 1) if pd.notna(x) else "—").tolist(),
            "KLPM":                  ta_df["klpm_v"].apply(
                lambda x: round(x, 2) if pd.notna(x) else "—").tolist(),
        })
        st.dataframe(disp, hide_index=True, use_container_width=True)
        df_download(disp, "t14_2")

        buf = io.BytesIO()
        disp.to_excel(buf, index=False, engine="openpyxl", sheet_name="TA_Finder")
        fname = (f"TA_Finder_{'+'.join(sel_omcs)}_{sel_metric.split()[0]}_"
                 f"{product}_{cy}_{period_lbl}.xlsx"
                 ).replace(" ", "_").replace("/", "").replace("(", "").replace(")", "")
        st.download_button("⬇ Download results as Excel (.xlsx)",
                           data=buf.getvalue(), file_name=fname,
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="fn_ta_dl")


# ─────────────────────────────────────────────────────────────────────────────
# ② Performance YoY bulk download
# ─────────────────────────────────────────────────────────────────────────────
def _perf_block(cy_g, ly_g, uset, ro_g, n_months, group_label, grp_name):
    """Return a list of row-dicts for one geographic group (OMC rows + subtotals)."""
    ind     = totals_row(cy_g, ly_g, uset)
    psu_row = totals_row(cy_g, ly_g, PSU) if uset == OMC_ORDER else ind
    cy_tot, ly_tot = ind["cy_tot"], ind["ly_tot"]
    ro_cnt  = (ro_g.groupby("omc")["sap_code"].nunique()
               .reindex(OMC_ORDER, fill_value=0).to_dict())
    tot_ros = sum(ro_cnt.get(o, 0) for o in uset)

    rows = []
    for omc in uset:
        n_ros = ro_cnt.get(omc, 0)
        r = {
            group_label:    grp_name,
            "OMC":           omc,
            "ROs":           n_ros,
            "RO Part%":      round(n_ros / tot_ros * 100, 1) if tot_ros else 0.0,
            "Vol CY (KL)":   round(ind[f"{omc}_cyvol"], 1),
            "Vol LY (KL)":   round(ind[f"{omc}_lyvol"], 1),
            "Change (KL)":   round(ind[f"{omc}_diffvol"], 1),
            "Growth %":      round(ind[f"{omc}_gr"], 2) if ind[f"{omc}_gr"] is not None else None,
            "KLPM CY":       round(klpm(ind[f"{omc}_cyvol"], n_months, n_ros), 2),
            "Share CY %":    round(ind[f"{omc}_cyshare"], 2),
            "Share LY %":    round(ind[f"{omc}_lyshare"], 2),
            "Share pp":      round(ind[f"{omc}_ppt"], 2),
            "Notional (KL)": round(ind[f"{omc}_notional"], 1),
        }
        if uset == OMC_ORDER and omc in PSU:
            r["PSU Share CY%"] = round(psu_row[f"{omc}_cyshare"], 2)
            r["PSU Share pp"]  = round(psu_row[f"{omc}_ppt"], 2)
        rows.append(r)

    for name, omcs, include in [
            ("Total PSU",  PSU,      True),
            ("Total PVT",  PVT,      uset == OMC_ORDER),
            ("Industry",   OMC_ORDER, uset == OMC_ORDER)]:
        if not include:
            continue
        cyv   = sum(ind[f"{o}_cyvol"] for o in omcs)
        lyv   = sum(ind[f"{o}_lyvol"] for o in omcs)
        cs    = cyv / cy_tot * 100 if cy_tot else 0.0
        ls    = lyv / ly_tot * 100 if ly_tot else 0.0
        gr    = (cyv - lyv) / lyv * 100 if lyv else None
        pp_   = cs - ls
        n_sub = sum(ro_cnt.get(o, 0) for o in omcs)
        r = {
            group_label:    grp_name,
            "OMC":           name,
            "ROs":           n_sub,
            "RO Part%":      round(n_sub / tot_ros * 100, 1) if tot_ros else 0.0,
            "Vol CY (KL)":   round(cyv, 1),
            "Vol LY (KL)":   round(lyv, 1),
            "Change (KL)":   round(cyv - lyv, 1),
            "Growth %":      round(gr, 2) if gr is not None else None,
            "KLPM CY":       round(klpm(cyv, n_months, n_sub), 2),
            "Share CY %":    round(cs, 2),
            "Share LY %":    round(ls, 2),
            "Share pp":      round(pp_, 2),
            "Notional (KL)": round(pp_ / 100 * cy_tot, 1) if cy_tot else 0.0,
        }
        if uset == OMC_ORDER:
            r["PSU Share CY%"] = None
            r["PSU Share pp"]  = None
        rows.append(r)

    rows.append({})   # blank separator row
    return rows


@st.fragment
def _render_perf_download(ctx):
    monthly    = ctx["monthly"]
    ro_master  = ctx["ro_master"]
    cy         = ctx["cy"]
    ly         = ctx["ly"]
    product    = ctx["product"]
    months     = ctx["months"]
    period_lbl = ctx["period_lbl"]
    n_months   = max(len(months), 1)

    if not ly:
        st.warning("No prior year in DB — CY vs LY report cannot be generated. "
                   "Select a later CY year in the sidebar.")
        return

    denom = st.radio("Denominator", ["Industry", "PSU"],
                     horizontal=True, key="pd_denom")
    uset  = OMC_ORDER if denom == "Industry" else PSU

    st.markdown("**Levels to include** (one sheet each)")
    lc1, lc2, lc3, lc4, lc5, lc6 = st.columns(6)
    inc_do   = lc1.checkbox("DO summary",     value=True,  key="pd_lv_do")
    inc_rsa  = lc2.checkbox("RSA",            value=True,  key="pd_lv_rsa")
    inc_dist = lc3.checkbox("District",       value=True,  key="pd_lv_dist")
    inc_com  = lc4.checkbox("COM",            value=True,  key="pd_lv_com")
    inc_hwyt = lc5.checkbox("Highway type",   value=False, key="pd_lv_hwyt")
    inc_hwno = lc6.checkbox("Highway number", value=False, key="pd_lv_hwno")

    st.caption("Scope filters — narrow the universe (empty = whole DO)")
    f_rsa, f_dist, f_com, f_hwytype, f_hwyno = _scope_widgets("pd", ctx)

    active = [(None,         "DO",          "DO",      inc_do),
              ("rsa_name",   "RSA",         "RSA",     inc_rsa),
              ("district",   "District",    "District",inc_dist),
              ("com",        "COM",         "COM",     inc_com),
              ("hwy_type",   "Highway Type","HwyType", inc_hwyt),
              ("highway_no", "Highway No.", "HwyNo",   inc_hwno)]
    active = [(col, label, sn) for col, label, sn, inc in active if inc]

    if not active:
        st.warning("Select at least one level.")
        return

    scope_m  = _apply_scope(monthly,   f_rsa, f_dist, f_com, f_hwytype, f_hwyno)
    scope_ro = _apply_scope(ro_master, f_rsa, f_dist, f_com, f_hwytype, f_hwyno)
    cy_f = scope_m[(scope_m["fy_code"] == cy) & (scope_m["product"] == product)
                   & (scope_m["month_index"].isin(months))]
    ly_f = scope_m[(scope_m["fy_code"] == ly) & (scope_m["product"] == product)
                   & (scope_m["month_index"].isin(months))]

    scope_tag = ("".join(
        [f"_RSA{len(f_rsa)}" if f_rsa else "",
         f"_Dist{len(f_dist)}" if f_dist else "",
         f"_COM{''.join(f_com)}" if f_com else ""])
    )
    fname = (f"Perf_YoY_{denom}_{product}_{cy}_{period_lbl}{scope_tag}.xlsx"
             ).replace(" ", "_").replace("/", "")

    if st.button("⬇ Generate & Download Performance YoY xlsx", key="pd_gen"):
        with st.spinner("Building xlsx…"):
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                from openpyxl.styles import Font
                from openpyxl.utils import get_column_letter
                for col, label, sname in active:
                    if col is None:
                        all_rows = _perf_block(
                            cy_f, ly_f, uset, scope_ro,
                            n_months, label, "Whole DO")
                    else:
                        all_rows = []
                        for grp in sorted(cy_f[col].dropna().unique()):
                            cy_g  = cy_f[cy_f[col] == grp]
                            ly_g  = ly_f[ly_f[col] == grp]
                            ro_g  = scope_ro[scope_ro[col] == grp]
                            all_rows.extend(
                                _perf_block(cy_g, ly_g, uset, ro_g,
                                            n_months, label, grp))
                    if not all_rows:
                        continue
                    # Fixed column order — sets are unordered so never use set here
                    ordered_cols = [label, "OMC", "ROs", "RO Part%",
                                    "Vol CY (KL)", "Vol LY (KL)", "Change (KL)",
                                    "Growth %", "KLPM CY",
                                    "Share CY %", "Share LY %", "Share pp",
                                    "Notional (KL)"]
                    if uset == OMC_ORDER:
                        ordered_cols += ["PSU Share CY%", "PSU Share pp"]
                    df_sh = pd.DataFrame(
                        [{k: r.get(k) for k in ordered_cols} for r in all_rows],
                        columns=ordered_cols)
                    df_sh.to_excel(writer, index=False, sheet_name=sname)
                    ws = writer.sheets[sname]
                    for cell in ws[1]:
                        cell.font = Font(bold=True)
                    for ri, r in enumerate(all_rows, start=2):
                        omc_v = r.get("OMC", "")
                        if str(omc_v).startswith("Total") or omc_v == "Industry":
                            for cell in ws[ri]:
                                cell.font = Font(bold=True)
                    for col_idx, col_cells in enumerate(ws.columns, start=1):
                        mx = max(len(str(c.value or "")) for c in col_cells)
                        ws.column_dimensions[
                            get_column_letter(col_idx)].width = min(mx + 2, 30)
        st.download_button(
            "⬇ Click to save",
            data=buf.getvalue(), file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="pd_save")
        st.caption(f"`{fname}` · sheets: {', '.join(s for _, _, s in active)}")


# ─────────────────────────────────────────────────────────────────────────────
# ③ TA Profile bulk download
# ─────────────────────────────────────────────────────────────────────────────
def _ta_rows_for_sheet(gdf, gtot):
    """Build (headers, data_rows) for one TA xlsx sheet (mirrors Tab 5 logic)."""
    OMC_SEQ = ["IOCL", "BPCL", "HPCL", "NEL", "RBML", "SIMPL"]

    def _blk(cy, ly, tcy, tly):
        diff = round(cy - ly, 1)
        gr   = f"{diff / ly * 100:.2f}%" if ly else "—"
        scy  = round(cy / tcy * 100, 2) if tcy else 0.0
        sly  = round(ly / tly * 100, 2) if tly else 0.0
        sgr  = round(scy - sly, 2)
        notv = round(sgr / 100 * tcy, 1)
        return round(cy, 1), round(ly, 1), diff, gr, scy, sly, sgr, notv

    hdrs = ["S.No", "Name of RO", "Location", "Oil Co",
            "MS CY (KL)", "MS LY (KL)", "MS +/-", "MS GR%",
            "HSD CY (KL)", "HSD LY (KL)", "HSD +/-", "HSD GR%",
            "MS Shr CY%", "MS Shr LY%", "MS ±pp", "MS Not(KL)",
            "HSD Shr CY%", "HSD Shr LY%", "HSD ±pp", "HSD Not(KL)"]

    gdf2 = gdf.copy()
    gdf2["tot_cy"] = gdf2["ms_cy"] + gdf2["hs_cy"]
    gdf2["_rank"] = gdf2["omc"].map(
        {o: i for i, o in enumerate(OMC_SEQ)}).fillna(99)
    gdf2 = gdf2.sort_values(["_rank", "tot_cy"], ascending=[True, False])

    data_rows, sn = [], 0
    for r in gdf2.itertuples():
        sn += 1
        ms = _blk(r.ms_cy, r.ms_ly, gtot["ms_cy"], gtot["ms_ly"])
        hs = _blk(r.hs_cy, r.hs_ly, gtot["hs_cy"], gtot["hs_ly"])
        data_rows.append([sn, r.ro, r.loc, r.omc] + list(ms) + list(hs))

    for omc in OMC_SEQ:
        sub = gdf2[gdf2["omc"] == omc]
        if sub.empty:
            continue
        ms = _blk(sub.ms_cy.sum(), sub.ms_ly.sum(), gtot["ms_cy"], gtot["ms_ly"])
        hs = _blk(sub.hs_cy.sum(), sub.hs_ly.sum(), gtot["hs_cy"], gtot["hs_ly"])
        data_rows.append(["", f"{omc} Sub Total",
                          f"Total ROs: {len(sub)}", ""] + list(ms) + list(hs))

    for label, grp in [("Total PSU",      ["IOCL","BPCL","HPCL"]),
                        ("Total Pvt.",     ["NEL","RBML","SIMPL"]),
                        ("Total Industry", OMC_SEQ)]:
        sub = gdf2[gdf2["omc"].isin(grp)]
        ms = _blk(sub.ms_cy.sum(), sub.ms_ly.sum(), gtot["ms_cy"], gtot["ms_ly"])
        hs = _blk(sub.hs_cy.sum(), sub.hs_ly.sum(), gtot["hs_cy"], gtot["hs_ly"])
        data_rows.append(["", label, "", ""] + list(ms) + list(hs))

    n_ro = max(len(gdf2), 1)
    ms = _blk(gtot["ms_cy"] / n_ro, gtot["ms_ly"] / n_ro,
               gtot["ms_cy"], gtot["ms_ly"])
    hs = _blk(gtot["hs_cy"] / n_ro, gtot["hs_ly"] / n_ro,
               gtot["hs_cy"], gtot["hs_ly"])
    data_rows.append(["", "TA Average", "", "Avg"] + list(ms) + list(hs))

    return hdrs, data_rows


@st.fragment
def _render_ta_bulk(ctx):
    monthly    = ctx["monthly"]
    ro_master  = ctx["ro_master"]
    cy         = ctx["cy"]
    ly         = ctx["ly"]
    product    = ctx["product"]
    months     = ctx["months"]
    period_lbl = ctx["period_lbl"]
    TA_NAME    = ctx["TA_NAME"]

    sel_mode = st.radio(
        "Which TAs to include",
        ["Top / Bottom N by criterion", "All TAs in filtered scope"],
        horizontal=True, key="tb_mode")

    selected_tas = []   # list of (ta_code, ta_name)

    if sel_mode == "Top / Bottom N by criterion":
        tc1, tc2, tc3 = st.columns([3, 2, 2])
        ta_metric = tc1.selectbox("Rank TAs by", [
            "IOCL share change (pp)",
            "IOCL volume CY (KL)",
            "IOCL volume change (KL)",
            "IOCL notional (KL)",
            "Industry volume CY (KL)",
        ], key="tb_metric")
        ta_dir = tc2.radio("Direction", ["Highest first", "Lowest first"],
                           horizontal=True, key="tb_dir")
        ta_n   = tc3.number_input("N", min_value=1, max_value=100,
                                  value=15, step=1, key="tb_n")
        denom_ta = st.radio("Ranking denominator", ["Industry", "PSU"],
                            horizontal=True, key="tb_denom")
        uset_ta  = OMC_ORDER if denom_ta == "Industry" else PSU
    else:
        denom_ta = "Industry"
        uset_ta  = OMC_ORDER

    st.caption("Scope filters — narrow which TAs are eligible (empty = whole DO)")
    f_rsa, f_dist, f_com, f_hwytype, f_hwyno = _scope_widgets("tb", ctx)

    scope_ro = _apply_scope(ro_master, f_rsa, f_dist, f_com, f_hwytype, f_hwyno)
    eligible  = set(scope_ro["ta_code"].dropna().unique())

    if sel_mode == "All TAs in filtered scope":
        selected_tas = [(t, TA_NAME.get(t, t)) for t in sorted(eligible)]
        if len(selected_tas) > 100:
            st.warning(
                f"**{len(selected_tas)} TAs** selected — this will create a large file. "
                "Consider switching to Top / Bottom N.")
    else:
        cy_pool = monthly[
            (monthly["fy_code"] == cy) & (monthly["product"] == product) &
            (monthly["month_index"].isin(months)) &
            (monthly["ta_code"].isin(eligible))
        ]
        ly_pool = (monthly[
            (monthly["fy_code"] == ly) & (monthly["product"] == product) &
            (monthly["month_index"].isin(months)) &
            (monthly["ta_code"].isin(eligible))
        ] if ly else monthly.iloc[0:0])

        if cy_pool.empty:
            st.info("No data for the selected scope and period.")
            return

        rsf = share_frame(cy_pool, ly_pool, ["ta_code"], uset_ta)
        if rsf.empty:
            st.info("No TA data available.")
            return

        mcol_map = {
            "IOCL share change (pp)":    "IOCL_ppt",
            "IOCL volume CY (KL)":       "IOCL_cyvol",
            "IOCL volume change (KL)":   "IOCL_diffvol",
            "IOCL notional (KL)":        "IOCL_notional",
            "Industry volume CY (KL)":   "cy_tot",
        }
        mcol = mcol_map[ta_metric]
        asc  = (ta_dir == "Lowest first")
        rsf  = rsf.sort_values(mcol, ascending=asc).head(int(ta_n))
        # share_frame returns reset index so ta_code is a column
        ta_col = "ta_code" if "ta_code" in rsf.columns else rsf.index.name
        ta_list = rsf["ta_code"].tolist() if "ta_code" in rsf.columns else rsf.index.tolist()
        selected_tas = [(str(t), TA_NAME.get(str(t), str(t))) for t in ta_list]

    if not selected_tas:
        st.info("No TAs match the current selection.")
        return

    n_ta = len(selected_tas)
    fname = (f"TA_Profile_Bulk_{n_ta}TAs_{product}_{cy}_{period_lbl}.xlsx"
             ).replace(" ", "_").replace("/", "")
    st.caption(f"Ready: **{n_ta} TA(s)** · one sheet per TA · `{fname}`")

    if st.button(f"⬇ Generate & Download {n_ta} TA Profile(s) (.xlsx)", key="tb_gen"):
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        wb.remove(wb.active)

        with st.spinner(f"Building {n_ta} TA sheet(s)…"):
            generated = 0
            for ta_code, ta_name in selected_tas:
                gdf, gtot = ta_volume_grid(monthly, ta_code, months, cy, ly)
                if gdf.empty:
                    continue
                hdrs, data_rows = _ta_rows_for_sheet(gdf, gtot)

                # Safe sheet name: strip openpyxl-forbidden chars, max 31 chars, unique
                _raw = (ta_name or ta_code) or ta_code
                for _ch in r'\/*?:[]':
                    _raw = _raw.replace(_ch, '_')
                sname = _raw[:28].strip('_ ') or ta_code[:10]
                existing = {ws.title for ws in wb.worksheets}
                base, sfx = sname, 1
                while sname in existing:
                    sname = f"{base[:25]}_{sfx}"; sfx += 1

                ws = wb.create_sheet(title=sname)
                title = (f"{ta_code} — {ta_name}  |  "
                         f"{product}  |  {cy} vs {ly or '—'}  |  {period_lbl}")
                ws.append([title])
                ws.merge_cells(start_row=1, start_column=1,
                                end_row=1, end_column=len(hdrs))
                ws.cell(1, 1).font = Font(bold=True, size=11)
                ws.append(hdrs)
                for cell in ws[2]:
                    cell.font = Font(bold=True)
                for ri, row in enumerate(data_rows, start=3):
                    ws.append(row)
                    if row and isinstance(row[0], str) and row[0] == "":
                        for cell in ws[ri]:
                            cell.font = Font(bold=True)
                for col_idx, col_cells in enumerate(ws.columns, start=1):
                    mx = max(len(str(c.value or "")) for c in col_cells)
                    ws.column_dimensions[
                        get_column_letter(col_idx)].width = min(mx + 2, 35)
                generated += 1

        buf = io.BytesIO()
        wb.save(buf)
        st.download_button(
            "⬇ Click to save",
            data=buf.getvalue(), file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="tb_save")
        st.success(f"Generated **{generated}** TA sheet(s).")


# ─────────────────────────────────────────────────────────────────────────────
# Main render
# ─────────────────────────────────────────────────────────────────────────────
def render(ctx):
    st.caption(
        "All three sections use the sidebar's **product**, **CY year**, and **period**. "
        "Each section has its own independent scope filters and does not affect "
        "the sidebar or other tabs."
    )

    with st.expander("① RO / TA Finder — personalised ranked query", expanded=True):
        _render_finder(ctx)

    with st.expander(
            "② Performance YoY Report — multi-level bulk xlsx download",
            expanded=False):
        _render_perf_download(ctx)

    with st.expander(
            "③ TA Profile Bulk Download — multiple TAs, one sheet per TA",
            expanded=False):
        _render_ta_bulk(ctx)

    with st.expander("④ RO Location & Contacts (IOCL geo master)", expanded=False):
        _render_geo(ctx)


def _render_geo(ctx):
    """Geo + contacts from dim_ro_geo (IOCL dealer master with GPS)."""
    from core import load_geo
    geo = load_geo()
    if geo.empty:
        st.info("Geo master not loaded.")
        return
    q = st.text_input("Search by SAP code or RO name", key="geo_q",
                      placeholder="e.g. 135645 or SWAMI SAMARTH")
    view = geo
    ro_master = ctx["ro_master"]
    names = ro_master.set_index("sap_code")["ro_name"].to_dict()
    geo = geo.assign(ro_name=geo.sap_code.map(names).fillna(""))
    if q and q.strip():
        qq = q.strip().upper()
        view = geo[geo.sap_code.str.contains(qq, na=False)
                   | geo.ro_name.str.upper().str.contains(qq, na=False)]
    else:
        view = geo
    if view.empty:
        st.warning("No match."); return
    if len(view) == 1 or (q and len(view) <= 5):
        for _, r in view.iterrows():
            st.markdown(
                f"**{r.ro_name or r.sap_code}** ({r.sap_code}) — {r.sales_area or ''}, "
                f"{r.divisional_office or ''}\n\n"
                f"📍 {r.address or '—'} · PIN {r.pin_code or '—'} · "
                f"Hwy {r.highway_no or '—'}\n\n"
                f"📞 Dealer: {r.primary_mobile or '—'} · ✉️ {r.email or '—'}\n\n"
                f"👤 Sales Officer: {r.sales_officer or '—'} ({r.so_mobile or '—'}) · "
                f"DRSM: {r.drsm_name or '—'}\n\n"
                f"🚚 Last load: {r.last_load_date or '—'}")
            if pd.notna(r.latitude) and pd.notna(r.longitude):
                st.map(pd.DataFrame({"lat": [r.latitude], "lon": [r.longitude]}),
                       zoom=12)
    else:
        show = view[["sap_code", "ro_name", "sales_area", "divisional_office",
                     "ro_type", "highway_no", "primary_mobile", "sales_officer",
                     "last_load_date"]]
        st.dataframe(show, hide_index=True, use_container_width=True, height=380)
        df_download(show, "geo_contacts")
        pts = view.dropna(subset=["latitude", "longitude"])
        if not pts.empty and st.checkbox("Show all on map", key="geo_map_all"):
            st.map(pts.rename(columns={"latitude": "lat", "longitude": "lon"})
                   [["lat", "lon"]], zoom=8)
